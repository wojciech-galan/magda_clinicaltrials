import argparse
import os
from typing import List
from typing import Tuple
from typing import Dict
from collections import namedtuple
from openpyxl.workbook import Workbook

Data = namedtuple('Data', 'study_id locations')
Location = namedtuple('Location', 'location country')


def read_tsv(path: str) -> List[Data]:
    res = []
    with open(path) as f:
        info_line = f.readline()
        splitted_info_line = info_line.split('\t')
        study_id_field_index = splitted_info_line.index('NCT Number')
        locations_field_index = splitted_info_line.index('Locations')
        for line in f:
            splitted_line = line.split('\t')
            res.append(Data(splitted_line[study_id_field_index], parse_locations(splitted_line[locations_field_index])))
    return res


def parse_locations(input_str: str) -> List[Location]:
    return [parse_location(x) for x in input_str.split('|')]


def parse_location(input_str: str) -> Location:
    if ', Korea, Republic of' in input_str:
        return Location(input_str.replace(', Korea, Republic of', ''), 'Korea, Republic of')
    else:
        splitted = input_str.rsplit(', ', 1)
        try:
            return Location(splitted[0], splitted[1])
        except IndexError:
            return Location('', '')


def write_study_locations(data: List[Data], path: str):
    wb = Workbook()
    ws = wb.active
    for i, (study_id, locations) in enumerate(data, 1):
        ws.cell(row=1, column=i * 2 - 1, value=study_id)
        ws.merge_cells(start_row=1, start_column=i * 2 - 1, end_row=1, end_column=i * 2)
        for j, (location, country) in enumerate(locations, 2):
            ws.cell(row=j, column=i * 2 - 1, value=location)
            ws.cell(row=j, column=i * 2, value=country)
    wb.save(path)


def write_location_studies(data: List[Data], path: str):
    location_country, location_study_id = rev_data(data)
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value='Institution')
    ws.cell(row=1, column=2, value='Country')
    ws.cell(row=1, column=3, value='Study ID')
    k = 2
    for location in location_country.keys():
        start_row = k
        ws.cell(row=k, column=1, value=location)
        ws.cell(row=k, column=2, value=location_country[location])
        for study_id in location_study_id[location]:
            ws.cell(row=k, column=3, value=study_id)
            k += 1
        if k - 1 > start_row:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=k - 1, end_column=1)
            ws.merge_cells(start_row=start_row, start_column=2, end_row=k - 1, end_column=2)
        k += 1
    wb.save(path)


def rev_data(data: List[Data]) -> Tuple[Dict[str, str]]:
    location_country = {}
    for study_id, locations in data:
        for location in locations:
            location_country[location.location] = location[1]
    location_study_id = {}
    for study_id, locations in data:
        for location in locations:
            try:
                location_study_id[location.location].append(study_id)
            except KeyError:
                location_study_id[location.location] = [study_id]
    return location_country, location_study_id


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='converts tab-separated data from clinicltrials.gov')
    parser.add_argument('infile', type=str, help='Input path')
    parser.add_argument('outfile', type=str, help='output path')
    parser.add_argument('--analysis', type=str, help='Type of analysis to be performed',
                        choices=['study-locations', 'location-studies'])
    parsed_args = parser.parse_args()
    infile = os.path.abspath(os.path.expanduser(parsed_args.infile))
    assert os.path.exists(infile)
    outfile = os.path.abspath(os.path.expanduser(parsed_args.outfile))
    assert os.path.exists(os.path.dirname(infile))
    data = read_tsv(infile)
    # write_study_locations(data, outfile)
    write_location_studies(data, outfile)
